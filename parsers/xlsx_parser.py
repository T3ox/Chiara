import openpyxl
import logging
from config.settings import TRUNCATION_LIMITS
from models.datatypes import FileContext
from .base import BaseParser

logger = logging.getLogger(__name__)

class XlsxParser(BaseParser):
    def parse(self, context: FileContext) -> FileContext:
        try:
            wb = openpyxl.load_workbook(context.file_path, data_only=True, read_only=True)
            output = f"Workbook: {context.file_name}\n"
            
            sheets = wb.sheetnames[:TRUNCATION_LIMITS["XLSX_MAX_SHEETS"]]
            for sheet_name in sheets:
                output += f"\n[SHEET] {sheet_name}\n"
                ws = wb[sheet_name]
                
                for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                    if row_idx >= TRUNCATION_LIMITS["XLSX_MAX_ROWS_PER_SHEET"]:
                        break
                    row_data = [str(cell) for cell in row if cell is not None]
                    if row_data:
                        output += " | ".join(row_data) + "\n"
                        
            context.extracted_text = output[:TRUNCATION_LIMITS["MAX_TEXT_CHARS"]]
            wb.close()
        except Exception as e:
            logger.error(f"Error parsing XLSX {context.file_name}: {e}")
            
        return context
