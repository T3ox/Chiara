import os
import base64
import io
from typing import Dict, Optional, Tuple
import fitz # PyMuPDF
from docx import Document
import openpyxl
from PIL import Image
from constants_py import TRUNCATION_LIMITS

class FileExtractor:
    @staticmethod
    def extract_text_from_docx(file_path: str, snippet: bool = False) -> str:
        doc = Document(file_path)
        full_text = []
        for para in doc.paragraphs:
            full_text.append(para.text)
        text = "\n".join(full_text)
        
        if snippet:
            return text[:TRUNCATION_LIMITS["DOCX_SNIPPET_CHARS"]]
        return text[:TRUNCATION_LIMITS["MAX_TEXT_CHARS"]]

    @staticmethod
    def extract_text_from_xlsx(file_path: str) -> str:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        output = f"Workbook: {os.path.basename(file_path)}\n"
        
        sheet_names = wb.sheetnames[:TRUNCATION_LIMITS["XLSX_MAX_SHEETS"]]
        for sheet_name in sheet_names:
            output += f"\n[SHEET] {sheet_name}\n"
            sheet = wb[sheet_name]
            
            rows_processed = 0
            for row in sheet.iter_rows(values_only=True):
                if rows_processed >= TRUNCATION_LIMITS["XLSX_MAX_ROWS_PER_SHEET"]:
                    break
                row_str = " | ".join([str(cell) if cell is not None else "" for cell in row])
                if row_str.strip():
                    output += row_str + "\n"
                    rows_processed += 1
                    
        return output[:TRUNCATION_LIMITS["MAX_TEXT_CHARS"]]

    @staticmethod
    def get_pdf_page_count(file_path: str) -> int:
        doc = fitz.open(file_path)
        return doc.page_count

    @staticmethod
    def get_pdf_thumbnail(file_path: str) -> str:
        doc = fitz.open(file_path)
        page = doc.load_page(0) # first page
        pix = page.get_pixmap()
        img_data = pix.tobytes("jpeg")
        return base64.b64encode(img_data).decode('utf-8')

    @staticmethod
    def resize_image(file_path: str) -> str:
        with Image.open(file_path) as img:
            # Convert to RGB if needed (e.g. for RGBA or CMYK)
            if img.mode in ("RGBA", "P"):
                img = img.convert("RGB")
                
            width, height = img.size
            max_side = TRUNCATION_LIMITS["IMAGE_MAX_SIDE"]
            
            if width > height:
                if width > max_side:
                    height = int(height * (max_side / width))
                    width = max_side
            else:
                if height > max_side:
                    width = int(width * (max_side / height))
                    height = max_side
                    
            img = img.resize((width, height), Image.Resampling.LANCZOS)
            buffered = io.BytesIO()
            img.save(buffered, format="JPEG", quality=80)
            return base64.b64encode(buffered.getvalue()).decode('utf-8')

    @staticmethod
    def get_pdf_data(file_path: str) -> str:
        with open(file_path, "rb") as f:
            return base64.b64encode(f.read()).decode('utf-8')
